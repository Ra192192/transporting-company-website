import asyncio
import logging
import os
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import psycopg
from psycopg.rows import dict_row
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, Update
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)


logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

DATABASE_URL = os.environ["DATABASE_URL"]
BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
POLL_INTERVAL_SECONDS = int(os.getenv("POLL_INTERVAL_SECONDS", "5"))

ADD_ADMIN_BUTTON = "Добавить админа"
SHOW_PENDING_BUTTON = "Показать ожидающие заявки"
EXPORT_ORDERS_BUTTON = "Выгрузить таблицу заказов"

ADMIN_KEYBOARD = ReplyKeyboardMarkup(
    [[ADD_ADMIN_BUTTON], [SHOW_PENDING_BUTTON], [EXPORT_ORDERS_BUTTON]],
    resize_keyboard=True,
    input_field_placeholder="Выберите действие",
)


def run_db(query: str, params: tuple = (), fetch: str | None = None):
    with psycopg.connect(DATABASE_URL, row_factory=dict_row) as conn:
        with conn.cursor() as cur:
            cur.execute(query, params)
            if fetch == "one":
                return cur.fetchone()
            if fetch == "all":
                return cur.fetchall()
        conn.commit()
    return None


def init_db() -> None:
    run_db(
        """
        CREATE TABLE IF NOT EXISTS telegram_admins (
            telegram_id BIGINT PRIMARY KEY,
            name VARCHAR(120) NOT NULL
        )
        """
    )
    run_db(
        """
        CREATE TABLE IF NOT EXISTS bot_state (
            key VARCHAR(80) PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    row = run_db("SELECT MAX(id) AS max_id FROM leads", fetch="one")
    max_id = row["max_id"] or 0
    run_db(
        """
        INSERT INTO bot_state (key, value)
        VALUES ('last_notified_lead_id', %s)
        ON CONFLICT (key) DO NOTHING
        """,
        (str(max_id),),
    )


def get_admins() -> list[dict]:
    return run_db(
        "SELECT telegram_id, name FROM telegram_admins ORDER BY name",
        fetch="all",
    )


def is_admin(telegram_id: int) -> bool:
    row = run_db(
        "SELECT 1 FROM telegram_admins WHERE telegram_id = %s",
        (telegram_id,),
        fetch="one",
    )
    return row is not None


def admin_count() -> int:
    row = run_db("SELECT COUNT(*) AS count FROM telegram_admins", fetch="one")
    return row["count"]


def add_admin(telegram_id: int, name: str) -> None:
    run_db(
        """
        INSERT INTO telegram_admins (telegram_id, name)
        VALUES (%s, %s)
        ON CONFLICT (telegram_id) DO UPDATE SET name = EXCLUDED.name
        """,
        (telegram_id, name),
    )


def get_last_notified_id() -> int:
    row = run_db(
        "SELECT value FROM bot_state WHERE key = 'last_notified_lead_id'",
        fetch="one",
    )
    return int(row["value"]) if row else 0


def set_last_notified_id(lead_id: int) -> None:
    run_db(
        """
        INSERT INTO bot_state (key, value)
        VALUES ('last_notified_lead_id', %s)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
        """,
        (str(lead_id),),
    )


def get_new_leads(after_id: int) -> list[dict]:
    return run_db(
        """
        SELECT id, name, phone, what_to_remove, status
        FROM leads
        WHERE id > %s
        ORDER BY id
        """,
        (after_id,),
        fetch="all",
    )


def get_unprocessed_leads() -> list[dict]:
    return run_db(
        """
        SELECT id, name, phone, what_to_remove, status
        FROM leads
        WHERE status = 'UNPROCESSED'
        ORDER BY id
        """,
        fetch="all",
    )


def get_all_leads() -> list[dict]:
    return run_db(
        """
        SELECT id, name, phone, what_to_remove, status
        FROM leads
        ORDER BY id
        """,
        fetch="all",
    )


def mark_lead_processed(lead_id: int) -> bool:
    row = run_db(
        """
        UPDATE leads
        SET status = 'PROCESSED'
        WHERE id = %s AND status = 'UNPROCESSED'
        RETURNING id
        """,
        (lead_id,),
        fetch="one",
    )
    return row is not None


def format_lead(lead: dict, title: str = "Заявка") -> str:
    description = lead["what_to_remove"] or "Не указано"
    return (
        f"{title}\n\n"
        f"ID: {lead['id']}\n"
        f"Имя: {lead['name']}\n"
        f"Телефон: {lead['phone']}\n"
        f"Что нужно вывезти: {description}\n"
        f"Статус: {lead['status']}"
    )


def process_keyboard(lead_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("Пометить как обработанную", callback_data=f"process:{lead_id}")]]
    )


def add_admin_help_text() -> str:
    return (
        "Добавить администратора:\n"
        "/add_admin <айди телеграмм аккаунта> <имя администратора>\n\n"
        "<телеграмм айди> привязано к аккаунту Telegram и можно получить через ботов, например @getmyid_bot.\n"
        "Имя выбираете сами.\n\n"
        "Пример:\n"
        "/add_admin 123456789 Иван"
    )


def build_orders_workbook(leads: list[dict]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заявки"

    headers = ["ID", "Имя", "Телефон", "Что нужно вывезти", "Статус"]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="EF1F2D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for lead in leads:
        sheet.append([
            lead["id"],
            lead["name"],
            lead["phone"],
            lead["what_to_remove"] or "",
            lead["status"],
        ])

    widths = [10, 24, 22, 46, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


async def is_allowed_user(update: Update) -> bool:
    user = update.effective_user
    if user is None:
        return False
    return admin_count() == 0 or is_admin(user.id)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or update.message is None:
        return

    if admin_count() == 0:
        name = user.full_name or user.username or str(user.id)
        add_admin(user.id, name)
        await update.message.reply_text(
            "Вы стали первым админом этого бота.",
            reply_markup=ADMIN_KEYBOARD,
        )
        return

    if not is_admin(user.id):
        return

    await update.message.reply_text(
        "Бот работает. Новые заявки будут приходить сюда.",
        reply_markup=ADMIN_KEYBOARD,
    )


async def add_admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or update.message is None or not is_admin(user.id):
        return

    if len(context.args) < 2:
        await update.message.reply_text(add_admin_help_text(), reply_markup=ADMIN_KEYBOARD)
        return

    try:
        telegram_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Телеграмм айди должен быть числом.", reply_markup=ADMIN_KEYBOARD)
        return

    name = " ".join(context.args[1:]).strip()
    if not name:
        await update.message.reply_text("Укажите имя админа.", reply_markup=ADMIN_KEYBOARD)
        return

    add_admin(telegram_id, name)
    await update.message.reply_text(f"Админ {name} добавлен.", reply_markup=ADMIN_KEYBOARD)


async def show_pending_leads(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None or not await is_allowed_user(update):
        return

    leads = get_unprocessed_leads()
    if not leads:
        await update.message.reply_text("Ожидающих заявок нет.", reply_markup=ADMIN_KEYBOARD)
        return

    await update.message.reply_text(
        f"Ожидающие заявки: {len(leads)}",
        reply_markup=ADMIN_KEYBOARD,
    )
    for lead in leads:
        await update.message.reply_text(
            format_lead(lead, "Ожидающая заявка"),
            reply_markup=process_keyboard(lead["id"]),
        )


async def export_orders(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None or not await is_allowed_user(update):
        return

    leads = get_all_leads()
    file = build_orders_workbook(leads)
    await update.message.reply_document(
        document=file,
        filename="orders.xlsx",
        caption=f"Таблица заказов. Записей: {len(leads)}",
        reply_markup=ADMIN_KEYBOARD,
    )


async def handle_menu_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None or not await is_allowed_user(update):
        return

    text = (update.message.text or "").strip()
    if text == ADD_ADMIN_BUTTON or text.lower() == "обавить админа":
        await update.message.reply_text(add_admin_help_text(), reply_markup=ADMIN_KEYBOARD)
        return

    if text == SHOW_PENDING_BUTTON:
        await show_pending_leads(update, context)
        return

    if text == EXPORT_ORDERS_BUTTON:
        await export_orders(update, context)


async def mark_processed_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    user = update.effective_user
    if query is None or user is None or not is_admin(user.id):
        return

    await query.answer()
    data = query.data or ""
    if not data.startswith("process:"):
        return

    try:
        lead_id = int(data.split(":", 1)[1])
    except ValueError:
        await query.edit_message_text("Не получилось определить заявку.")
        return

    if mark_lead_processed(lead_id):
        await query.edit_message_text(f"Заявка #{lead_id} помечена как обработанная.")
    else:
        await query.edit_message_text(f"Заявка #{lead_id} уже обработана или не найдена.")


async def notify_admins(application: Application, admins: Iterable[dict], lead: dict) -> None:
    for admin in admins:
        try:
            await application.bot.send_message(
                chat_id=admin["telegram_id"],
                text=format_lead(lead, "Новая заявка"),
                reply_markup=process_keyboard(lead["id"]),
            )
        except Exception:
            logger.exception("Failed to notify admin %s", admin["telegram_id"])


async def poll_new_leads(application: Application) -> None:
    while True:
        try:
            admins = get_admins()
            if admins:
                last_id = get_last_notified_id()
                leads = get_new_leads(last_id)
                for lead in leads:
                    await notify_admins(application, admins, lead)
                    set_last_notified_id(lead["id"])
        except Exception:
            logger.exception("Lead polling failed")

        await asyncio.sleep(POLL_INTERVAL_SECONDS)


async def post_init(application: Application) -> None:
    init_db()
    asyncio.create_task(poll_new_leads(application))


def main() -> None:
    application = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("add_admin", add_admin_command))
    application.add_handler(CallbackQueryHandler(mark_processed_callback, pattern=r"^process:\d+$"))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu_button))
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
